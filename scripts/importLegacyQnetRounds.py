from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Iterable

import pdfplumber
from openpyxl import load_workbook
from pypdf import PdfReader


ROOT = Path(__file__).resolve().parents[1]
DATA_PATH = ROOT / "src/data/releasedExamQuestions.json"
QNET_DIR = ROOT / "source_pdfs/qnet"
ANSWER_DIR = ROOT / "tmp/qnet/answers_20_29"

CHOICE_SYMBOLS = "①②③④⑤"
CHOICE_TO_NUMBER = {symbol: index + 1 for index, symbol in enumerate(CHOICE_SYMBOLS)}


ROUND_SPECS = {
    26: {
        "year": 2015,
        "question_files": [
            {
                "path": QNET_DIR / "2015년 제26회 공인중개사 2차 A형.pdf",
                "offset": 0,
            },
        ],
    },
    29: {
        "year": 2018,
        "question_files": [
            {
                "path": QNET_DIR / "2018년 제29회 시험문제지_2차 1교시 A형.pdf",
                "offset": 0,
            },
            {
                "path": QNET_DIR / "2018년 제29회 시험문제지_2차 2교시 A형.pdf",
                "offset": 80,
            },
        ],
    },
}

ANSWER_29_A = {
    **dict(zip(range(1, 41), [4, 3, 5, 4, 2, 5, 2, 4, 1, 1, 5, 5, 2, 5, 3, 1, 3, 5, 3, 4, 3, 4, 5, 3, 5, 2, 1, 1, 2, 4, 3, 1, 2, 5, 2, 1, 4, 3, 4, 1])),
    **dict(zip(range(41, 81), [5, 4, 3, 5, 1, 1, 5, 2, 4, 3, 5, 3, 5, 1, 2, 4, 0, 2, 3, 5, 3, 5, 1, 2, 4, 1, 4, 3, 4, 5, 2, 5, 2, 3, 0, 3, 1, 2, 3, 4])),
    **dict(zip(range(81, 121), [1, 2, 4, 5, 3, 1, 3, 4, 2, 3, 5, 2, 3, 3, 4, 1, 2, 2, 4, 5, 1, 5, 4, 5, 1, 4, 4, 3, 4, 5, 1, 3, 2, 3, 5, 2, 1, 5, 5, 2])),
}


def clean_text(value: str) -> str:
    value = value.replace("\u200b", "")
    value = re.sub(r"[ \t]+", " ", value)
    value = re.sub(r"\n{3,}", "\n\n", value)
    return value.strip()


def line_is_noise(line: str) -> bool:
    compact = re.sub(r"\s+", "", line)
    if not compact:
        return True
    if re.search(r"공인중개사.*A형-\d+-\d+", compact):
        return True
    if re.search(r"A-\d+-\d+|\(2\)|차$", compact):
        return True
    if compact.startswith(("년도제회", "<유의사항", "문제지형별", "수험번호", "가답안")):
        return True
    if compact.startswith("○") and len(compact) < 45:
        return True
    return False


def extract_column_text(pdf_path: Path) -> str:
    chunks: list[str] = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            middle = page.width / 2
            for box in ((0, 0, middle, page.height), (middle, 0, page.width, page.height)):
                text = page.crop(box).extract_text(x_tolerance=1, y_tolerance=3) or ""
                lines = [line for line in text.splitlines() if not line_is_noise(line)]
                chunks.append("\n".join(lines))
    return "\n".join(chunks)


def split_question_blocks(text: str) -> list[tuple[int, str]]:
    starts = list(re.finditer(r"(?m)^(\d{1,3})\.\s*", text))
    blocks: list[tuple[int, str]] = []
    for index, match in enumerate(starts):
        end = starts[index + 1].start() if index + 1 < len(starts) else len(text)
        blocks.append((int(match.group(1)), text[match.start():end].strip()))
    return blocks


def parse_question_block(local_number: int, block: str) -> dict | None:
    block = re.sub(r"^\d{1,3}\.\s*", "", block).strip()
    parts = re.split(f"([{CHOICE_SYMBOLS}])", block)
    if len(parts) < 11:
        return None

    stem = clean_text(parts[0])
    choices: list[str] = []
    for index in range(1, len(parts), 2):
        symbol = parts[index]
        value = parts[index + 1] if index + 1 < len(parts) else ""
        if symbol in CHOICE_SYMBOLS:
            choices.append(clean_text(value))

    if len(choices) < 5:
        return None

    return {
        "localNumber": local_number,
        "questionText": stem,
        "choices": choices[:5],
    }


def subject_meta(round_number: int, global_number: int) -> tuple[str, str, int, int]:
    if 1 <= global_number <= 40:
        return "중개사법", "중개사법", global_number, global_number

    if round_number <= 26 and 41 <= global_number <= 80:
        local = global_number - 40
        if local <= 12:
            sub_subject = "지적법"
        elif local <= 24:
            sub_subject = "등기법"
        else:
            sub_subject = "세법"
        return "공시세법", sub_subject, local, local

    if round_number <= 26 and 81 <= global_number <= 120:
        return "공법", "공법", global_number - 80, global_number - 40

    if 41 <= global_number <= 80:
        return "공법", "공법", global_number - 40, global_number

    if 81 <= global_number <= 120:
        local = global_number - 80
        if local <= 12:
            sub_subject = "지적법"
        elif local <= 24:
            sub_subject = "등기법"
        else:
            sub_subject = "세법"
        return "공시세법", sub_subject, local, local
    raise ValueError(f"unsupported global question number: {global_number}")


def parse_answers_26() -> dict[int, int]:
    workbook = load_workbook(ANSWER_DIR / "26_1.xlsx", data_only=True)
    worksheet = workbook["Sheet1"]
    answers: dict[int, int] = {}
    in_a_type = False

    for row in worksheet.iter_rows(values_only=True):
        values = [value for value in row if value is not None]
        if not values:
            continue
        first = str(values[0])
        if "A형" in first:
            in_a_type = True
            continue
        if "B형" in first:
            break
        if not in_a_type:
            continue
        for index in range(0, len(row), 2):
            number = row[index] if index < len(row) else None
            answer = row[index + 1] if index + 1 < len(row) else None
            if isinstance(number, int) and isinstance(answer, int):
                answers[number] = answer

    return answers


def answer_map_for_round(round_number: int) -> dict[int, int]:
    if round_number == 26:
        return parse_answers_26()
    if round_number == 29:
        return ANSWER_29_A
    raise ValueError(f"unsupported round: {round_number}")


def law_ref_for(subject: str, sub_subject: str) -> str:
    if subject == "중개사법":
        return "공인중개사법령 및 부동산거래신고 등에 관한 법령"
    if subject == "공법":
        return "부동산공법"
    if sub_subject == "세법":
        return "부동산세법"
    if sub_subject == "등기법":
        return "부동산등기법"
    return "공간정보의 구축 및 관리 등에 관한 법령"


def id_prefix_for(subject: str, sub_subject: str) -> str:
    if subject == "중개사법":
        return "broker-law"
    if subject == "공법":
        return "public-law"
    return "tax-law"


def build_round_questions(round_number: int) -> list[dict]:
    spec = ROUND_SPECS[round_number]
    answer_map = answer_map_for_round(round_number)
    parsed_by_global: dict[int, dict] = {}

    for file_spec in spec["question_files"]:
        text = extract_column_text(file_spec["path"])
        for local_number, block in split_question_blocks(text):
            parsed = parse_question_block(local_number, block)
            if not parsed:
                continue
            global_number = local_number + file_spec["offset"]
            parsed_by_global[global_number] = parsed

    questions: list[dict] = []
    for global_number in range(1, 121):
        parsed = parsed_by_global.get(global_number)
        answer = answer_map.get(global_number)
        if answer is None:
            raise ValueError(f"round {round_number} question {global_number} is not importable")

        subject, sub_subject, exam_number, display_number = subject_meta(round_number, global_number)
        prefix = id_prefix_for(subject, sub_subject)
        source = "EBS/Q-Net 공개 기출문제 및 최종정답"
        needs_review = parsed is None or answer not in range(1, 6)
        question_text = parsed["questionText"] if parsed else f"제{round_number}회 {global_number}번 문항 원문 추출 검수 필요"
        choices = parsed["choices"] if parsed else [
            "원문 선택지 추출 검수 필요",
            "원문 선택지 추출 검수 필요",
            "원문 선택지 추출 검수 필요",
            "원문 선택지 추출 검수 필요",
            "원문 선택지 추출 검수 필요",
        ]
        stored_answer = answer if answer in range(1, 6) else 1
        questions.append({
            "id": f"real-{round_number}-{prefix}-{exam_number:02d}",
            "sourceRound": round_number,
            "sourceYear": spec["year"],
            "subject": subject,
            "subSubject": sub_subject,
            "examNumber": exam_number,
            "displayNumber": display_number,
            "chapter": sub_subject,
            "topic": question_text.split("\n", 1)[0][:80],
            "lawRef": law_ref_for(subject, sub_subject),
            "difficulty": "normal",
            "sourceType": "original",
            "category": "past",
            "frequencyScore": 90,
            "questionText": question_text,
            "choices": choices,
            "answer": stored_answer,
            "explanation": "최종정답이 전항정답으로 표시된 문항입니다. 앱 단일정답 구조 확장 전까지 검수 필요 상태로 보관합니다." if needs_review else f"정답은 {answer}번입니다. EBS/Q-Net 공개 기출문제 A형과 최종정답을 대조했습니다.",
            "lawUpdateNote": "현행법 보정 없음: 기출 원문과 보정 메타를 분리했습니다.",
            "sourceTitle": f"제{round_number}회 실제 기출",
            "originalSource": source,
            "isLawUpdated": False,
            "lawUpdateDescription": "",
            "needsReview": needs_review,
        })

    return questions


def sort_key(question: dict) -> tuple[int, int, int]:
    subject_order = {"중개사법": 1, "공법": 2, "공시세법": 3}
    return (question["sourceRound"], subject_order[question["subject"]], question["examNumber"])


def main() -> None:
    existing = json.loads(DATA_PATH.read_text())
    importing_rounds = set(ROUND_SPECS)
    kept = [question for question in existing if question["sourceRound"] not in importing_rounds]

    imported: list[dict] = []
    for round_number in sorted(importing_rounds):
        round_questions = build_round_questions(round_number)
        if len(round_questions) != 120:
            raise ValueError(f"round {round_number} imported {len(round_questions)} questions")
        imported.extend(round_questions)
        print(f"round {round_number}: imported {len(round_questions)} questions")

    merged = sorted([*kept, *imported], key=sort_key)
    DATA_PATH.write_text(json.dumps(merged, ensure_ascii=False, indent=2) + "\n")
    print(f"wrote {len(merged)} questions to {DATA_PATH}")


if __name__ == "__main__":
    main()
